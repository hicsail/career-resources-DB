import json
import re
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from urllib.parse import urlparse

# === Constants ===
EXCEL_FILE = "/Users/asadeg02/Downloads/Resource Glossary AB_AT.xlsx"
SHEET_NAME = "Sheet1"
START_ROW = 3
END_ROW = 115
TAG_COLUMN_INDEX = 1    # Column A
TITLE_COLUMN_INDEX = 2  # Column B
LINK_COLUMN_INDEX = 3   # Column C

TAG_DROPDOWNS_FILE_PATH = "tag_dropdowns.json"
RESOURCE_DICT_FILE_PATH = "resource_dict.json"

# === Helpers ===
def is_google_link(link):
    return isinstance(link, str) and ('drive.google.com' in link or 'docs.google.com' in link or 'forms.gle' in link)

def extract_hyperlink(ws: Worksheet, row: int, col: int) -> str:
    cell = ws.cell(row=row, column=col)
    return cell.hyperlink.target if cell.hyperlink else cell.value

# === Core Functions ===
def parse_tag_dropdowns(sheet: Worksheet) -> dict:
    tag_dropdowns = {}

    for dv in sheet.data_validations.dataValidation:
        if dv.type == "list" and dv.formula1:
            options = dv.formula1.strip('"').split(',')
            for cell_range in dv.ranges.ranges:
                # Only consider ranges that include TAG_COLUMN_INDEX
                if TAG_COLUMN_INDEX >= cell_range.min_col and TAG_COLUMN_INDEX <= cell_range.max_col:
                    for row in range(cell_range.min_row, cell_range.max_row + 1):
                        if START_ROW <= row <= END_ROW:
                            tag_value = sheet.cell(row=row, column=TAG_COLUMN_INDEX).value
                            if tag_value:
                                tag_dropdowns[tag_value] = options
    return tag_dropdowns

def parse_resources(ws: Worksheet, tag_dropdowns: dict) -> dict:
    resource_dict = {}

    for row in range(START_ROW, END_ROW + 1):
        title = ws.cell(row=row, column=TITLE_COLUMN_INDEX).value
        link = extract_hyperlink(ws, row, LINK_COLUMN_INDEX)
        tag_label = ws.cell(row=row, column=TAG_COLUMN_INDEX).value
        
        if title and link and is_google_link(link):
            tag_options = tag_dropdowns.get(tag_label, [])            
            resource_dict[title] = {
                "link": convert_to_direct_link(link),
                "tags": tag_options
            }

    return resource_dict

def save_json(data: dict, file_path: str):
    with open(file_path, "w") as f:
        json.dump(data, f, indent=2)

def extract_file_id(url: str) -> str | None:
    """
    Extracts the Google file ID from Docs or Drive URLs.
    """
    # Match document, spreadsheet, presentation, or drive file
    patterns = [
        r'document/d/([a-zA-Z0-9_-]+)',
        r'spreadsheets/d/([a-zA-Z0-9_-]+)',
        r'presentation/d/([a-zA-Z0-9_-]+)',
        r'file/d/([a-zA-Z0-9_-]+)'
    ]
    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            return match.group(1)
    return None

def is_google_docs_link(url: str) -> bool:
    return bool(re.search(r'docs\.google\.com/(document|spreadsheets|presentation)/d/', url))

def is_google_drive_file(url: str) -> bool:
    return bool(re.search(r'drive\.google\.com/file/d/', url))

def convert_to_direct_link(url: str) -> str | None:
    """
    Converts a Google Doc/Sheet/Slide or Drive link to a direct downloadable/viewable link.
    """
    file_id = extract_file_id(url)
    if not file_id:
        return None

    if is_google_docs_link(url):
        # For Docs/Sheets/Slides, export as PDF
        if "document" in url:
            return f"https://docs.google.com/document/d/{file_id}/export?format=pdf"
        elif "spreadsheets" in url:
            return f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"
        elif "presentation" in url:
            return f"https://docs.google.com/presentation/d/{file_id}/export/pdf"
    elif is_google_drive_file(url):
        # For Drive file (PDFs, images, zips, etc.)
        return f"https://drive.google.com/uc?export=download&id={file_id}"

    return None

# === Main Runner ===
def main():
    wb = load_workbook(EXCEL_FILE, data_only=True)
    sheet = wb[SHEET_NAME]

    tag_dropdowns = parse_tag_dropdowns(sheet)
    resources = parse_resources(sheet, tag_dropdowns)

    save_json(tag_dropdowns, TAG_DROPDOWNS_FILE_PATH)
    save_json(resources, RESOURCE_DICT_FILE_PATH)
    print(f"Saved {len(tag_dropdowns)} tag dropdown entries to '{TAG_DROPDOWNS_FILE}'")
    print(f"Saved {len(resources)} resources to '{RESOURCE_DICT_FILE}'")

if __name__ == "__main__":
    main()